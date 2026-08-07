"""export 모듈 테스트 (CSV / XLSX 내용 검증)."""

from __future__ import annotations

import csv

import pytest

from mdlogger import db, export
from mdlogger.db import COLUMNS


def make_conn_with_rows():
    conn = db.connect(":memory:")
    db.init_db(conn)
    db.insert_game(
        conn,
        dict(
            played_at="2026-06-19T10:00:00",
            result="win",
            turn_order="first",
            my_deck="스네이크아이",
            opp_deck="블루아이즈",
            turns=5,
            end_reason="regular",
            score_after=2600,
            note="좋음",
        ),
    )
    db.insert_game(
        conn,
        dict(
            played_at="2026-06-19T10:05:00",
            result="lose",
            turn_order="second",
            my_deck="티아라멘츠",
            opp_deck="엑조디아",
            turns=8,
            end_reason="surrender",
            score_after=1300,
            note="",
        ),
    )
    return conn


def test_export_csv(tmp_path):
    conn = make_conn_with_rows()
    out = tmp_path / "games.csv"
    export.export_csv(out, db.get_all_games(conn))

    with open(out, encoding="utf-8-sig") as f:
        reader = list(csv.reader(f))

    assert reader[0] == COLUMNS
    assert len(reader) == 3  # 헤더 + 2행
    record = dict(zip(COLUMNS, reader[1]))
    assert record["my_deck"] == "스네이크아이"
    assert record["opp_deck"] == "블루아이즈"
    assert record["result"] == "win"
    assert record["score_after"] == "2600"
    assert record["note"] == "좋음"


@pytest.mark.parametrize("prefix", ["=", "+", "-", "@", "\t", "\r", "\n"])
def test_spreadsheet_formula_prefixes_are_escaped(prefix):
    assert export._spreadsheet_safe(f"{prefix}danger") == f"'{prefix}danger"


def test_export_csv_escapes_formula(tmp_path):
    conn = make_conn_with_rows()
    db.update_game(
        conn,
        1,
        {
            "played_at": "2026-06-19T10:00:00",
            "result": "win",
            "turn_order": "first",
            "my_deck": "스네이크아이",
            "opp_deck": "=1+1",
            "turns": 5,
            "end_reason": "regular",
            "score_after": 2600,
            "note": "@SUM(1,1)",
        },
    )
    out = tmp_path / "games.csv"
    export.export_csv(out, db.get_all_games(conn))

    with open(out, encoding="utf-8-sig") as f:
        record = dict(zip(COLUMNS, list(csv.reader(f))[1]))

    assert record["opp_deck"] == "'=1+1"
    assert record["note"] == "'@SUM(1,1)"


def test_export_xlsx(tmp_path):
    from openpyxl import load_workbook

    conn = make_conn_with_rows()
    out = tmp_path / "games.xlsx"
    export.export_xlsx(out, db.get_all_games(conn))

    wb = load_workbook(out)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))

    assert list(values[0]) == COLUMNS
    assert len(values) == 3
    assert values[1][COLUMNS.index("my_deck")] == "스네이크아이"
    assert values[1][COLUMNS.index("opp_deck")] == "블루아이즈"
    assert values[1][COLUMNS.index("score_after")] == 2600  # 정수 유지
    assert values[2][COLUMNS.index("end_reason")] == "surrender"


def test_export_xlsx_escapes_formula(tmp_path):
    from openpyxl import load_workbook

    conn = make_conn_with_rows()
    db.update_game(
        conn,
        1,
        {
            "played_at": "2026-06-19T10:00:00",
            "result": "win",
            "turn_order": "first",
            "my_deck": "스네이크아이",
            "opp_deck": "=1+1",
            "turns": 5,
            "end_reason": "regular",
            "score_after": 2600,
            "note": "+danger",
        },
    )
    out = tmp_path / "games.xlsx"
    export.export_xlsx(out, db.get_all_games(conn))

    ws = load_workbook(out, data_only=False).active
    opp_deck = ws.cell(row=2, column=COLUMNS.index("opp_deck") + 1)
    note = ws.cell(row=2, column=COLUMNS.index("note") + 1)
    assert opp_deck.value == "'=1+1"
    assert opp_deck.data_type == "s"
    assert note.value == "'+danger"
    assert note.data_type == "s"
