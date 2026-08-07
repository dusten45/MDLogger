"""현재 게임 동작과 새 service/repository 경계의 characterization test."""

from __future__ import annotations

import csv
import sqlite3

import pytest

from mdlogger.db import COLUMNS
from mdlogger.game_service import GameService, SqliteGameRepository


def sample(**over) -> dict:
    record = {
        "played_at": "2026-06-19T10:00:00",
        "result": "win",
        "turn_order": "first",
        "my_deck": "스네이크아이",
        "opp_deck": "블루아이즈",
        "turns": 5,
        "end_reason": "regular",
        "score_after": 2600,
        "note": "",
    }
    record.update(over)
    return record


def make_service() -> GameService:
    return GameService.open(":memory:")


def test_service_preserves_record_workflow_and_statistics():
    games = make_service()

    first_id = games.insert_game(sample())
    second_id = games.insert_game(
        sample(
            played_at="2026-06-19T10:05:00",
            result="lose",
            turn_order="second",
            my_deck="티아라멘츠",
            opp_deck="엑조디아",
            turns=8,
            end_reason="surrender",
            score_after=1300,
        )
    )

    assert (first_id, second_id) == (1, 2)
    assert games.count_games() == 2
    assert games.get_last_score() == 1300
    assert games.get_last_my_deck() == "티아라멘츠"
    last_game = games.get_last_game()
    assert last_game is not None
    assert last_game["id"] == second_id
    assert [row["id"] for row in games.get_all_games()] == [first_id, second_id]

    summary = games.get_summary()
    assert summary["total"] == 2
    assert summary["wins"] == 1
    assert summary["losses"] == 1
    assert summary["first_winrate"] == 100.0
    assert summary["second_winrate"] == 0.0
    assert summary["avg_turns"] == 6.5
    assert games.get_deck_matchups("second") == [
        {
            "deck": "엑조디아",
            "games": 1,
            "wins": 0,
            "losses": 1,
            "winrate": 0.0,
        }
    ]

    games.update_game(
        first_id,
        sample(
            played_at="2026-06-19T10:00:00",
            result="lose",
            opp_deck="엔디미온",
            score_after=2000,
            note="편집됨",
        ),
    )
    edited = games.get_game(first_id)
    assert edited is not None
    assert edited["played_at"] == "2026-06-19T10:00:00"
    assert edited["opp_deck"] == "엔디미온"
    assert edited["note"] == "편집됨"

    games.delete_game(second_id)
    assert games.count_games() == 1
    last_game = games.get_last_game()
    assert last_game is not None
    assert last_game["id"] == first_id


def test_service_exports_current_repository_rows(tmp_path):
    games = make_service()
    games.insert_game(sample(note="내보내기"))

    csv_path = tmp_path / "games.csv"
    xlsx_path = tmp_path / "games.xlsx"
    games.export_csv(csv_path)
    games.export_xlsx(xlsx_path)

    with csv_path.open(encoding="utf-8-sig") as file:
        csv_rows = list(csv.reader(file))
    assert csv_rows[0] == COLUMNS
    assert dict(zip(COLUMNS, csv_rows[1]))["note"] == "내보내기"

    from openpyxl import load_workbook

    worksheet = load_workbook(xlsx_path).active
    assert worksheet is not None
    assert (
        worksheet.cell(row=2, column=COLUMNS.index("opp_deck") + 1).value
        == "블루아이즈"
    )


def test_service_owns_injected_repository_lifecycle():
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    repository = SqliteGameRepository(connection)
    games = GameService(repository)

    connection.execute("CREATE TABLE marker (value INTEGER)")
    games.close()

    with pytest.raises(sqlite3.ProgrammingError, match="closed database"):
        connection.execute("SELECT * FROM marker")
